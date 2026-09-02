"""
Report Service — generates an Excel daily report of all job activity.

Uses openpyxl to create a formatted spreadsheet with:
- Summary sheet: daily stats
- Jobs sheet: all jobs found today with scores
- Applications sheet: emails sent
"""
import logging
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.config.settings import REPORTS_DIR

logger = logging.getLogger(__name__)


HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="F0F4FF")
GREEN_FILL = PatternFill("solid", fgColor="C8F7C5")
YELLOW_FILL = PatternFill("solid", fgColor="FFF9C4")
RED_FILL = PatternFill("solid", fgColor="FFCCCC")

thin = Side(style="thin", color="CCCCCC")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)


class ReportService:
    """
    Generates formatted Excel reports of daily job search activity.
    """

    def __init__(self):
        self.reports_dir = Path(REPORTS_DIR)
        self.reports_dir.mkdir(parents=True, exist_ok=True)

    def generate_daily_report(
        self,
        scored_jobs: list[dict],
        applications: list[dict],
        run_stats: dict,
    ) -> str:
        """
        Generate a comprehensive daily Excel report.

        Args:
            scored_jobs: [{"job_id", "title", "company", "score", "reasoning", "job_url", ...}]
            applications: [{"title", "company", "to_email", "email_sent", "resume_path"}]
            run_stats: {"total_fetched", "new_jobs", "qualified", "emails_sent", ...}

        Returns:
            Path to generated Excel file
        """
        today = date.today().strftime("%Y-%m-%d")
        filename = f"JobReport_{today}.xlsx"
        filepath = self.reports_dir / filename

        wb = openpyxl.Workbook()

        # Build sheets
        self._build_summary_sheet(wb, today, run_stats, scored_jobs)
        self._build_jobs_sheet(wb, scored_jobs)
        self._build_applications_sheet(wb, applications)

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        try:
            wb.save(str(filepath))
        except PermissionError:
            from datetime import datetime as _dt
            suffix = _dt.now().strftime("%H%M%S")
            filename = f"JobReport_{today}_{suffix}.xlsx"
            filepath = self.reports_dir / filename
            wb.save(str(filepath))

        logger.info(f"📊 Excel report saved: {filepath}")
        return str(filepath)

    def _build_summary_sheet(self, wb, today: str, stats: dict, scored_jobs: list[dict]):
        ws = wb.create_sheet("📊 Summary", 0)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        # Title
        ws.merge_cells("A1:B1")
        ws["A1"] = f"🤖 AI Job Agent — Daily Report — {today}"
        ws["A1"].font = Font(bold=True, size=14, color="1a1a2e")
        ws["A1"].alignment = Alignment(horizontal="center")

        rows = [
            ("Metric", "Value"),
            ("Date", today),
            ("Total Jobs Fetched", stats.get("total_fetched", 0)),
            ("New Jobs Saved", stats.get("new_jobs", 0)),
            ("Duplicates Skipped", stats.get("duplicates_skipped", 0)),
            ("Jobs Scored (AI 0-2 Yr)", len(scored_jobs)),
            ("Qualified Jobs (≥65)", stats.get("qualified", 0)),
            ("Tailored ATS Resumes Made", stats.get("resumes_generated", 0)),
            ("Posts Scanned for Recruiter Emails", stats.get("posts_scanned_for_hr", 0)),
            ("Recruiter Emails Discovered", stats.get("recruiter_emails_found", 0)),
            ("Recruiter Outreach Sent (Stream A)", stats.get("emails_sent", 0)),
            ("Direct Apply Links Prepared (Stream B)", stats.get("direct_applied", 0)),
            ("WhatsApp Match Alerts Sent", "Yes" if stats.get("whatsapp_sent") else "No"),
        ]

        for i, (key, val) in enumerate(rows, start=3):
            ws[f"A{i}"] = key
            ws[f"B{i}"] = val
            ws[f"A{i}"].border = thin_border
            ws[f"B{i}"].border = thin_border
            if i == 3:  # Header row
                ws[f"A{i}"].fill = HEADER_FILL
                ws[f"A{i}"].font = HEADER_FONT
                ws[f"B{i}"].fill = HEADER_FILL
                ws[f"B{i}"].font = HEADER_FONT
            elif i % 2 == 0:
                ws[f"A{i}"].fill = ALT_FILL
                ws[f"B{i}"].fill = ALT_FILL

        # All Eligible matches table
        eligible_jobs = [j for j in scored_jobs if j.get("score", 0) >= 65]
        if not eligible_jobs:
            eligible_jobs = sorted(scored_jobs, key=lambda x: x.get("score", 0), reverse=True)
        else:
            eligible_jobs = sorted(eligible_jobs, key=lambda x: x.get("score", 0), reverse=True)

        ws["A17"] = f"🏆 All Eligible Job Matches ({len(eligible_jobs)} Roles with 1-Click Apply Links)"
        ws["A17"].font = Font(bold=True, size=12)

        headers = ["Rank", "Title", "Company", "Score", "Action", "Apply Link"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=18, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = thin_border

        for rank, job in enumerate(eligible_jobs, 1):
            row = 18 + rank
            score = job.get("score", 0)
            fill = GREEN_FILL if score >= 75 else YELLOW_FILL if score >= 65 else RED_FILL
            for col, val in enumerate([
                rank,
                job.get("title", ""),
                job.get("company", ""),
                f"{score}/100",
                job.get("recommended_action", ""),
                job.get("job_url", ""),
            ], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                if col in (1, 4):
                    cell.fill = fill

        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["F"].width = 50

    def _build_jobs_sheet(self, wb, scored_jobs: list[dict]):
        ws = wb.create_sheet("📋 All Scored Jobs")

        headers = [
            "Job ID", "Title", "Company", "Score", "Recommended Action",
            "Reasoning", "Matching Skills", "Missing Skills", "URL",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = thin_border

        sorted_jobs = sorted(scored_jobs, key=lambda x: x.get("score", 0), reverse=True)

        for row_num, job in enumerate(sorted_jobs, start=2):
            score = job.get("score", 0)
            fill = GREEN_FILL if score >= 75 else YELLOW_FILL if score >= 65 else RED_FILL

            values = [
                job.get("job_id", ""),
                job.get("title", ""),
                job.get("company", ""),
                score,
                job.get("recommended_action", ""),
                job.get("reasoning", ""),
                ", ".join(job.get("matching_skills", [])),
                ", ".join(job.get("missing_skills", [])),
                job.get("job_url", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                if col == 4:  # Score column gets color
                    cell.fill = fill
                elif row_num % 2 == 0:
                    cell.fill = ALT_FILL

        # Column widths
        widths = [8, 35, 25, 8, 12, 50, 35, 35, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _build_applications_sheet(self, wb, applications: list[dict]):
        ws = wb.create_sheet("📧 Recruiter Outreach & Applications")

        headers = [
            "Role Title", "Company", "Score", "Route", "Recruiter Email Discovered",
            "Outreach Email Sent", "Tailored Resume Attached", "Apply Link", "Date",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = thin_border

        for row_num, app in enumerate(applications, start=2):
            values = [
                app.get("title", ""),
                app.get("company", ""),
                app.get("score", ""),
                app.get("route", "Direct Application Link"),
                app.get("to_email", "None (Direct Link)"),
                "✅ SENT" if app.get("email_sent") else "Prepared (Link)",
                "✅ Attached" if app.get("resume_path") else "❌",
                app.get("job_url", ""),
                app.get("date", date.today().strftime("%Y-%m-%d")),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = ALT_FILL

        widths = [35, 25, 8, 22, 32, 18, 22, 45, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def generate_recruiter_report(
        self,
        recruiter_entries: list[dict],
        stats: dict,
    ) -> str:
        """
        Generate a dedicated Excel report specifically for Recruiter & HR Email Outreach.
        """
        today = date.today().strftime("%Y-%m-%d")
        filename = f"Recruiter_HR_Report_{today}.xlsx"
        filepath = self.reports_dir / filename

        wb = openpyxl.Workbook()

        # Sheet 1: Main Post Feed & Outreach Report
        ws_main = wb.create_sheet("📬 HR Posts & Outreach", 0)
        
        # Title Banner
        ws_main.merge_cells("A1:H1")
        ws_main["A1"] = f"📬 LinkedIn Recruiter & HR Outreach Report — {today}"
        ws_main["A1"].font = Font(bold=True, size=15, color="1a1a2e")
        ws_main["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_main.row_dimensions[1].height = 28

        # KPI Summary Table
        ws_main["A3"] = "Metric"
        ws_main["B3"] = "Value"
        ws_main["A3"].fill = HEADER_FILL
        ws_main["A3"].font = HEADER_FONT
        ws_main["B3"].fill = HEADER_FILL
        ws_main["B3"].font = HEADER_FONT
        ws_main["A3"].border = thin_border
        ws_main["B3"].border = thin_border

        kpi_rows = [
            ("Report Date", today),
            ("Candidate Profile", "Kamal Kumar (0-2 Yrs / 1+ Yrs)"),
            ("Total Posts Scanned", stats.get("posts_scanned_for_hr", len(recruiter_entries))),
            ("Verified Recruiter Emails", stats.get("recruiter_emails_found", len([e for e in recruiter_entries if e.get("hr_email")]))),
            ("Cold Outreach Dispatched", stats.get("emails_sent", 0)),
            ("Direct Post Links Prepared", stats.get("direct_applied", len(recruiter_entries))),
        ]

        for i, (k, v) in enumerate(kpi_rows, start=4):
            ws_main[f"A{i}"] = k
            ws_main[f"B{i}"] = v
            ws_main[f"A{i}"].border = thin_border
            ws_main[f"B{i}"].border = thin_border
            if i % 2 == 0:
                ws_main[f"A{i}"].fill = ALT_FILL
                ws_main[f"B{i}"].fill = ALT_FILL

        # Table Header for Posts & Outreach
        start_row = 11
        ws_main.merge_cells(f"A{start_row-1}:H{start_row-1}")
        ws_main[f"A{start_row-1}"] = "📋 Discovered LinkedIn Hiring Posts & Outreach Summary"
        ws_main[f"A{start_row-1}"].font = Font(bold=True, size=12, color="1a1a2e")

        headers = [
            "#", "Role Title", "Company / Poster", "Match Score",
            "Recruiter / HR Email", "LinkedIn Post Link (1-Click)", "Outreach Status", "Tailored Resume"
        ]

        for col, h in enumerate(headers, 1):
            cell = ws_main.cell(row=start_row, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_main.row_dimensions[start_row].height = 22

        current_row = start_row + 1
        link_font = Font(color="0000FF", underline="single")

        for idx, entry in enumerate(recruiter_entries, start=1):
            sent = entry.get("email_sent", False)
            hr_email = entry.get("hr_email")
            status_str = "✅ EMAILED HR" if sent else ("Verified HR (Pending)" if hr_email else "Post Link Prepared")
            post_url = entry.get("job_url", "")

            ws_main.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws_main.cell(row=current_row, column=2, value=entry.get("title", "Backend Developer"))
            ws_main.cell(row=current_row, column=3, value=entry.get("company", "LinkedIn Recruiter"))
            ws_main.cell(row=current_row, column=4, value=entry.get("score", "90/100")).alignment = Alignment(horizontal="center")
            ws_main.cell(row=current_row, column=5, value=hr_email or "Not in Post Text")
            
            # Clickable Post Link
            link_cell = ws_main.cell(row=current_row, column=6)
            if post_url and post_url.startswith("http"):
                link_cell.value = f'=HYPERLINK("{post_url}", "🔗 View LinkedIn Post")'
                link_cell.font = link_font
            else:
                link_cell.value = post_url or "https://www.linkedin.com/feed/"

            ws_main.cell(row=current_row, column=7, value=status_str).alignment = Alignment(horizontal="center")
            ws_main.cell(row=current_row, column=8, value="✅ Attached (PDF)" if entry.get("resume_path") else "—").alignment = Alignment(horizontal="center")

            for col in range(1, 9):
                c = ws_main.cell(row=current_row, column=col)
                c.border = thin_border
                if col == 7 and sent:
                    c.fill = GREEN_FILL
                elif current_row % 2 == 0:
                    c.fill = ALT_FILL

            ws_main.row_dimensions[current_row].height = 20
            current_row += 1

        widths = [6, 32, 28, 14, 32, 28, 22, 18]
        for i, w in enumerate(widths, 1):
            ws_main.column_dimensions[get_column_letter(i)].width = w

        # Sheet 2: Detailed Email Log
        ws_log = wb.create_sheet("📧 Detailed Outreach Log")
        log_headers = [
            "#", "Role Title", "Company", "Recruiter Email", "Post URL",
            "Status", "Subject", "Date & Timestamp"
        ]
        for col, h in enumerate(log_headers, 1):
            cell = ws_log.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = thin_border

        for idx, entry in enumerate(recruiter_entries, start=2):
            sent = entry.get("email_sent", False)
            status_str = "✅ SENT" if sent else ("Verified Email" if entry.get("hr_email") else "Post Link")
            post_url = entry.get("job_url", "")
            
            ws_log.cell(row=idx, column=1, value=idx - 1)
            ws_log.cell(row=idx, column=2, value=entry.get("title", ""))
            ws_log.cell(row=idx, column=3, value=entry.get("company", ""))
            ws_log.cell(row=idx, column=4, value=entry.get("hr_email") or "—")
            
            lcell = ws_log.cell(row=idx, column=5)
            if post_url and post_url.startswith("http"):
                lcell.value = f'=HYPERLINK("{post_url}", "{post_url[:40]}...")'
                lcell.font = link_font
            else:
                lcell.value = post_url

            ws_log.cell(row=idx, column=6, value=status_str)
            ws_log.cell(row=idx, column=7, value=entry.get("subject", ""))
            ws_log.cell(row=idx, column=8, value=entry.get("date", datetime.now().strftime("%Y-%m-%d %H:%M")))

            for col in range(1, 9):
                c = ws_log.cell(row=idx, column=col)
                c.border = thin_border
                if idx % 2 == 0:
                    c.fill = ALT_FILL

        log_widths = [6, 30, 25, 30, 45, 18, 40, 20]
        for i, w in enumerate(log_widths, 1):
            ws_log.column_dimensions[get_column_letter(i)].width = w

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        try:
            wb.save(str(filepath))
        except PermissionError:
            suffix = datetime.now().strftime("%H%M%S")
            filename = f"Recruiter_HR_Report_{today}_{suffix}.xlsx"
            filepath = self.reports_dir / filename
            wb.save(str(filepath))

        logger.info(f"📊 Recruiter & HR report saved: {filepath}")
        return str(filepath)
